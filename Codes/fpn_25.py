# -*- coding: utf-8 -*-


import os, cv2
import numpy as np
import pandas as pd
import random, tqdm
import seaborn as sns
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings("ignore")
import torch
import torch.nn as nn
from torch.utils.data import DataLoader
import albumentations as album
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image
import segmentation_models_pytorch as smp
from segmentation_models_pytorch import utils as smp_utils
from itertools import product
import gc
DATA_DIR = ''


torch.cuda.empty_cache()
metadata_df_normal = pd.read_csv(os.path.join(DATA_DIR, 'metadata_normal.csv'))
metadata_df_malignant = pd.read_csv(os.path.join(DATA_DIR, 'metadata_malignant.csv'))
metadata_df_benign = pd.read_csv(os.path.join(DATA_DIR, 'metadata_benign.csv'))
metadata_df= pd.concat([metadata_df_normal,metadata_df_malignant,metadata_df_benign], ignore_index=True)

metadata_df = metadata_df[['Name', 'Image_Path', 'Mask_Path']]
print(metadata_df)
metadata_df['Image_Path'] = metadata_df['Image_Path'].apply(lambda img_pth: os.path.join(DATA_DIR, img_pth))
metadata_df['Mask_Path'] = metadata_df['Mask_Path'].apply(lambda img_pth: os.path.join(DATA_DIR, img_pth))
print(metadata_df)
metadata_df = metadata_df.sample(frac=1).reset_index(drop=True)

# Perform 90/10 split for train / val
test_df = metadata_df.sample(frac=0.2, random_state=42)
train_df = metadata_df.drop(test_df.index)
valid_df= metadata_df.sample(frac=(1/8), random_state=42)
train_df= metadata_df.drop(valid_df.index)
                    
print(metadata_df, valid_df, train_df)

class_dict = pd.read_csv(os.path.join(DATA_DIR, 'labels_class_dict_2.csv'))
# Get class names
print(class_dict)
class_names = class_dict['class_names'].tolist()
# Get class RGB values
class_rgb_values = class_dict[['r','g','b']].values.tolist()
print(class_rgb_values)
select_classes = ['tumor', 'non_tumor']


select_class_indices = [class_names.index(cls.lower()) for cls in select_classes]
select_class_rgb_values =  np.array(class_rgb_values)[select_class_indices]
# helper function for data visualization
def visualize(**images):
    """
    Plot images in one row
    """
    n_images = len(images)
    plt.figure(figsize=(20,8))
    for idx, (name, image) in enumerate(images.items()):
        plt.subplot(1, n_images, idx + 1)
        plt.xticks([]); 
        plt.yticks([])
        # get title from the parameter names
        plt.title(name.replace('_',' ').title(), fontsize=20)
        plt.imshow(image)
    plt.show()

# Perform one hot encoding on label
def one_hot_encode(label, label_values):
    """
    Convert a segmentation image label array to one-hot format
    by replacing each pixel value with a vector of length num_classes
    # Arguments
        label: The 2D array segmentation image label
        label_values
        
    # Returns
        A 2D array with the same width and hieght as the input, but
        with a depth size of num_classes
    """
    semantic_map = []
    for colour in label_values:
        equality = np.equal(label, colour)
        class_map = np.all(equality, axis = -1)
        semantic_map.append(class_map)
    semantic_map = np.stack(semantic_map, axis=-1)

    return semantic_map
    
# Perform reverse one-hot-encoding on labels / preds
def reverse_one_hot(image):
    """
    Transform a 2D array in one-hot format (depth is num_classes),
    to a 2D array with only 1 channel, where each pixel value is
    the classified class key.
    # Arguments
        image: The one-hot format image 
        
    # Returns
        A 2D array with the same width and hieght as the input, but
        with a depth size of 1, where each pixel value is the classified 
        class key.
    """
    x = np.argmax(image, axis = -1)
    return x

# Perform colour coding on the reverse-one-hot outputs
def colour_code_segmentation(image, label_values):
    """
    Given a 1-channel array of class keys, colour code the segmentation results.
    # Arguments
        image: single channel array where each value represents the class key.
        label_values

    # Returns
        Colour coded image for segmentation visualization
    """
    colour_codes = np.array(label_values)
    x = colour_codes[image.astype(int)]

    return x
class BackgroundDataset(torch.utils.data.Dataset):

    """
    
    Args:
        df (str): DataFrame containing images / labels paths
        class_rgb_values (list): RGB values of select classes to extract from segmentation mask
        augmentation (albumentations.Compose): data transfromation pipeline 
            (e.g. flip, scale, etc.)
        preprocessing (albumentations.Compose): data preprocessing 
            (e.g. noralization, shape manipulation, etc.)
    
    """
    def __init__(
            self, 
            df,
            class_rgb_values=None, 
            augmentation=None, 
            preprocessing=None,
    ):
        self.image_paths = df['Image_Path'].tolist()
        self.mask_paths = df['Mask_Path'].tolist()
        
        self.class_rgb_values = class_rgb_values
        self.augmentation = augmentation
        self.preprocessing = preprocessing
    
    def __getitem__(self, i):
        
        # read images and masks
        image = cv2.cvtColor(cv2.imread(self.image_paths[i]), cv2.COLOR_BGR2RGB)
        mask = cv2.cvtColor(cv2.imread(self.mask_paths[i]), cv2.COLOR_BGR2RGB)
          
        # one-hot-encode the mask
        mask = one_hot_encode(mask, self.class_rgb_values).astype('float')
        
        # apply augmentations
        if self.augmentation:
            sample = self.augmentation(image=image, mask=mask)
            image, mask = sample['image'], sample['mask']
        
        # apply preprocessing
        if self.preprocessing:
            sample = self.preprocessing(image=image, mask=mask)
            image, mask = sample['image'], sample['mask']
            
        return image, mask
        
    def __len__(self):
        # return length of 
        return len(self.image_paths)

dataset = BackgroundDataset(train_df, class_rgb_values=select_class_rgb_values)
random_idx = random.randint(0, len(dataset)-1)
image, mask = dataset[2]

visualize(
    original_image = image,
    ground_truth_mask = colour_code_segmentation(reverse_one_hot(mask), select_class_rgb_values),
    one_hot_encoded_mask = reverse_one_hot(mask)
)


parameters={
    'back_bone': ['resnet50', 'resnext50_32x4d',  ]
}

parameters_combinations = list(product(*parameters.values()))
print(len(parameters_combinations))

results_df = pd.DataFrame(columns=list(parameters.keys())+['epoch'] + ['iou_score'] +['dice_loss'])
name='results.xlsx'
results_df.to_excel(name)

best_iou_score = 0.0
train_logs, valid_logs=[],[]
for i, params in enumerate(parameters_combinations):
  back_bone = params[0]

  def get_training_augmentation():
      train_transform = [
          album.OneOf(
              [
                  album.HorizontalFlip(p=1),
                  album.VerticalFlip(p=1),
                  album.RandomRotate90(p=1),
              ],
              p=0.5,
          ),
          album.GaussNoise(p=0.5),
          # Histogram equalization
          album.Equalize(p=0.7),
          album.Resize(width=512, height=512),
      ]
      return album.Compose(train_transform)


  def get_validation_augmentation():
      # Add sufficient padding to ensure image is divisible by 32
      test_transform = [
          album.PadIfNeeded(min_height=512, min_width=512, always_apply=True, border_mode=0),
          album.Resize(width=512, height=512),
      ]
      return album.Compose(test_transform)



  def to_tensor(x, **kwargs):
      return x.transpose(2, 0, 1).astype('float32')

  def get_preprocessing(preprocessing_fn=None):
      """Construct preprocessing transform    
      Args:
          preprocessing_fn (callable): data normalization function 
              (can be specific for each pretrained neural network)
      Return:
          transform: albumentations.Compose
      """   
      _transform = []
      if preprocessing_fn:
          _transform.append(album.Lambda(image=preprocessing_fn))
      _transform.append(album.Lambda(image=to_tensor, mask=to_tensor))
          
      return album.Compose(_transform)


  augmented_dataset = BackgroundDataset(
      train_df, 
      augmentation=get_training_augmentation(),
      class_rgb_values=select_class_rgb_values,
  )

  random_idx = random.randint(0, len(augmented_dataset)-1)

  # Different augmentations on image/mask pairs
  for idx in range(3):
      image, mask = augmented_dataset[idx]
      visualize(
          original_image = image,
          ground_truth_mask = colour_code_segmentation(reverse_one_hot(mask), select_class_rgb_values),
          one_hot_encoded_mask = reverse_one_hot(mask)
      )

  ENCODER = back_bone
  ENCODER_WEIGHTS = 'imagenet'
  CLASSES = select_classes
  ACTIVATION = 'sigmoid' # could be None for logits or 'softmax2d' for multiclass segmentation

  # create segmentation model with pretrained encoder
  if back_bone=='xception':
    model = smp.FPN(
        encoder_name=ENCODER, 
        encoder_depth=65,
        encoder_weights=ENCODER_WEIGHTS, 
        classes=len(CLASSES), 
        activation=ACTIVATION,
    )
  else:
    model = smp.FPN(
        encoder_name=ENCODER, 
        encoder_weights=ENCODER_WEIGHTS, 
        classes=len(CLASSES), 
        activation=ACTIVATION,
    )

  preprocessing_fn = smp.encoders.get_preprocessing_fn(ENCODER, ENCODER_WEIGHTS)

  # Get train and val dataset instances
  train_dataset = BackgroundDataset(
      train_df, 
      augmentation=get_training_augmentation(),
      preprocessing=get_preprocessing(preprocessing_fn),
      class_rgb_values=select_class_rgb_values,
  )

  valid_dataset = BackgroundDataset(
      valid_df, 
      augmentation=get_validation_augmentation(), 
      preprocessing=get_preprocessing(preprocessing_fn),
      class_rgb_values=select_class_rgb_values,
  )

  # Get train and val data loaders
  train_loader = DataLoader(train_dataset, batch_size=16, shuffle=True, num_workers=0)
  valid_loader = DataLoader(valid_dataset, batch_size=16, shuffle=False, num_workers=0)

  # Set flag to train the model or not. If set to 'False', only prediction is performed (using an older model checkpoint)
  TRAINING = True

  # Set num of epochs
  EPOCHS = 25

  # Set device: `cuda` or `cpu`
  DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

  gc.collect()
  torch.cuda.empty_cache()
  # define loss function
  loss = smp.losses.DiceLoss('multilabel')
  loss.__name__= 'dice_loss'

  # define metrics
  metrics = [
      smp_utils.metrics.IoU(threshold=0.5),
  ]

  # define optimizer
  optimizer = torch.optim.Adam([ 
      dict(params=model.parameters(), lr=0.0001),
  ])

  # define learning rate scheduler (not used in this NB)
  lr_scheduler = torch.optim.lr_scheduler.CosineAnnealingWarmRestarts(
      optimizer, T_0=1, T_mult=2, eta_min=5e-5,
  )

  # load best saved model checkpoint from previous commit (if present)
  if os.path.exists('../input/pyramid-scene-parsing-pspnet-resnext50-pytorch/best_model.pth'):
      model = torch.load('../input/pyramid-scene-parsing-pspnet-resnext50-pytorch/best_model.pth', map_location=DEVICE)
  train_epoch = smp.utils.train.TrainEpoch(
      model,
      loss=loss,
      metrics=metrics, 
      optimizer=optimizer,
      device=DEVICE,
      verbose=True,
  )

  valid_epoch = smp.utils.train.ValidEpoch(
      model, 
      loss=loss, 
      metrics=metrics, 
      device=DEVICE,
      verbose=True,
  )

  if TRAINING:

      
      train_logs_list, valid_logs_list = [], []

      for epoch in range(0, EPOCHS):

          # Perform training & validation
          print('\n',{*params},'Epoch: ', {epoch})
          gc.collect()
          torch.cuda.empty_cache()
          train_logs = train_epoch.run(train_loader)
          valid_logs = valid_epoch.run(valid_loader)
          print(train_logs, valid_logs)
          train_logs_list.append(train_logs)
          valid_logs_list.append(valid_logs)
          rn=len(results_df)
          results_df.loc[rn] = [*params, epoch, train_logs['iou_score'], train_logs['dice_loss'] ]

          # Save model if a better val IoU score is obtained
          if best_iou_score < valid_logs['iou_score']:
              best_iou_score = valid_logs['iou_score']
              torch.save(model, './best_model.pth')
              print('Model saved!')
          

# load best saved model checkpoint from the current run
if os.path.exists('./best_model.pth'):
    best_model = torch.load('./best_model.pth', map_location=DEVICE)
    print('Loaded PSPNet model from this run.')

# load best saved model checkpoint from previous commit (if present)
elif os.path.exists('../input/pyramid-scene-parsing-pspnet-resnext50-pytorch/best_model.pth'):
    best_model = torch.load('../input/pyramid-scene-parsing-pspnet-resnext50-pytorch/best_model.pth', map_location=DEVICE)
    print('Loaded PSPNet model from a previous commit.')

# create test dataloader (with preprocessing operation: to_tensor(...))
name='results.xlsx'
results_df.to_excel(name)

test_dataset = BackgroundDataset(
    test_df, 
    augmentation=get_validation_augmentation(), 
    preprocessing=get_preprocessing(preprocessing_fn),
    class_rgb_values=select_class_rgb_values,
)

test_dataloader = DataLoader(test_dataset)

# test dataset for visualization (without preprocessing augmentations & transformations)
test_dataset_vis = BackgroundDataset(
    test_df,
    class_rgb_values=select_class_rgb_values,
    augmentation=get_validation_augmentation(),
)

# get a random test image/mask index
random_idx = random.randint(0, len(test_dataset_vis)-1)
image, mask = test_dataset_vis[random_idx]

visualize(
    original_image = image,
    ground_truth_mask = colour_code_segmentation(reverse_one_hot(mask), select_class_rgb_values),
    one_hot_encoded_mask = reverse_one_hot(mask)
)

# Center crop padded image / mask to original image dims
def crop_image(image, true_dimensions):
    return album.CenterCrop(p=1, height=true_dimensions[0], width=true_dimensions[1])(image=image)
sample_preds_folder = 'sample_predictions/'
if not os.path.exists(sample_preds_folder):
    os.makedirs(sample_preds_folder)


def IoU(pred_mask, gt_mask):

    # Threshold predicted mask
    pred_mask = (pred_mask > 0.5).astype(np.uint8)
    
    
    # Calculate intersection and union
    intersection = np.sum(pred_mask * gt_mask)
    union = np.sum((pred_mask + gt_mask) > 0)
    
    # Calculate IoU score
    if union == 0:
        iou_score = 0
    else:
        iou_score = intersection / union
    
    return iou_score
    return iou_score
    
def DiceLoss(pred, target):
    intersection = np.logical_and(target, pred)
    pred = (pred > 0.5).astype(np.float32)
    target = target.astype(np.float32)
    intersection = np.sum(pred * target)
    union = np.sum(pred) + np.sum(target)
    dice_loss = 1 - ((2 * intersection) / (union + intersection))
    return dice_loss

def calculate_dice_iou(pred_mask, gt_mask):
    # convert masks to binary format
    threshold=0.5
    pred_mask = (pred_mask > threshold).astype(np.uint8)
    gt_mask = (gt_mask > threshold).astype(np.uint8)

    # calculate true positive (TP), false positive (FP), and false negative (FN) values
    TP = np.sum(pred_mask * gt_mask)
    FP = np.sum(pred_mask * (1 - gt_mask))
    FN = np.sum((1 - pred_mask) * gt_mask)

    # calculate Dice Loss and IoU score
    dice_loss = 1 - ((2 * TP) / ((2 * TP )+ FP + FN))
    iou_score = TP / (TP + FP + FN)

    return dice_loss, iou_score

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Define the table headers and add them to the worksheet
headers = ['Index', 'Image', 'Ground Truth Mask', 'Predicted Mask', 'IoU Score', 'Dice Loss']
ws.append(headers)
row_num = 2
for idx in range(len(test_dataset)):

    image, gt_mask = test_dataset[idx]
    image_vis = test_dataset_vis[idx][0].astype('uint8')
    true_dimensions = image_vis.shape
    x_tensor = torch.from_numpy(image).to(DEVICE).unsqueeze(0)
    # Predict test image
    pred_mask = best_model(x_tensor)
    pred_mask = pred_mask.detach().squeeze().cpu().numpy()
    # Convert pred_mask from `CHW` format to `HWC` format
    pred_mask = np.transpose(pred_mask,(1,2,0))
    # Get prediction channel corresponding to foreground
    #pred_foreground_heatmap = crop_image(pred_mask[:,:,select_classes.index('foreground')], true_dimensions)['image']
    pred_mask = crop_image(colour_code_segmentation(reverse_one_hot(pred_mask), select_class_rgb_values), true_dimensions)['image']
    # Convert gt_mask from `CHW` format to `HWC` format
    gt_mask = np.transpose(gt_mask,(1,2,0))
    gt_mask = crop_image(colour_code_segmentation(reverse_one_hot(gt_mask), select_class_rgb_values), true_dimensions)['image']
    cv2.imwrite(os.path.join(sample_preds_folder, f"sample_pred_{idx}.png"), np.hstack([image_vis, gt_mask, pred_mask])[:,:,::-1])
    
    visualize(
        original_image = image_vis,
        ground_truth_mask = gt_mask,
        predicted_mask = pred_mask,
    )


    # Convert images to `Image` objects
    gt_mask = gt_mask.astype('uint8')
    pred_mask= pred_mask.astype('uint8')
    img1 = Image.fromarray(image_vis)
    img2 = Image.fromarray(gt_mask)
    img3 = Image.fromarray(pred_mask)

    ws.cell(row=row_num, column=1).value=idx
    # Add images to cells in the table
    col_num = 2
    img_dict={'image_vis':img1, 'gt_mask':img2, 'pred_mask':img3}
    for img in img_dict:
        col_letter = get_column_letter(col_num)
        img_path = os.path.join(sample_preds_folder, f"sample_pred_{idx}_{img}.png")
        img_dict[img].save(img_path)
        ws.add_image(openpyxl.drawing.image.Image(img_path), f"{col_letter}{row_num}")
        col_num += 1
    # Add metrics to remaining cells in the row



    dice_loss, iou_score= calculate_dice_iou(pred_mask, gt_mask)
    for metric in [iou_score, dice_loss]:
        ws.cell(row=row_num, column=col_num).value = metric
        col_num += 1
    # Increment the row number for the next iteration
    row_num += 1

    
    
    
wb.save('predictions.xlsx')
    
test_epoch = smp.utils.train.ValidEpoch(
    model,
    loss=loss, 
    metrics=metrics, 
    device=DEVICE,
    verbose=True,
)

valid_logs = test_epoch.run(test_dataloader)
print("Evaluation on Test Data: ")
print(f"Mean IoU Score: {valid_logs['iou_score']:.4f}")
print(f"Mean Dice Loss: {valid_logs['dice_loss']:.4f}")

train_logs_df = pd.DataFrame(train_logs_list)
valid_logs_df = pd.DataFrame(valid_logs_list)
train_logs_df.T

plt.figure(figsize=(20,8))
plt.plot(train_logs_df.index.tolist(), train_logs_df.iou_score.tolist(), lw=3, label = 'Train')
plt.plot(valid_logs_df.index.tolist(), valid_logs_df.iou_score.tolist(), lw=3, label = 'Valid')
plt.xlabel('Epochs', fontsize=21)
plt.ylabel('IoU Score', fontsize=21)
plt.title('IoU Score Plot', fontsize=21)
plt.legend(loc='best', fontsize=16)
plt.grid()
plt.savefig('iou_score_plot.png')
plt.show()


plt.figure(figsize=(20,8))
plt.plot(train_logs_df.index.tolist(), train_logs_df.dice_loss.tolist(), lw=3, label = 'Train')
plt.plot(valid_logs_df.index.tolist(), valid_logs_df.dice_loss.tolist(), lw=3, label = 'Valid')
plt.xlabel('Epochs', fontsize=21)
plt.ylabel('Dice Loss', fontsize=21)
plt.title('Dice Loss Plot', fontsize=21)
plt.legend(loc='best', fontsize=16)
plt.grid()
plt.savefig('dice_loss_plot.png')
plt.show()